import openpyxl
from openpyxl.styles.colors import Color

INPUT_FILE = '/Users/zhaoxue/Documents/beads/fuse-bead-tool/Colors.xlsx'
OUTPUT_FILE = '/Users/zhaoxue/Documents/beads/fuse-bead-tool/Colors-new.xlsx'

max_row = 218

def hex_to_rgb(hex_color):
    r = int(hex_color[1:3], 16)
    g = int(hex_color[3:5], 16)
    b = int(hex_color[5:7], 16)
    return f"{r << 24 | g << 16 | b << 8 | 255}"

# Open excel file
wb = openpyxl.load_workbook(INPUT_FILE)
sheet = wb.active

for row_num in range(2, max_row + 1):
    color_value = sheet.cell(row=row_num, column=1).value
    if color_value:
        # Set background
        color_obj = Color(rgb=f'00{color_value}')
        for col_num in range(1,5):
            sheet.cell(row=row_num, column=col_num).fill = openpyxl.styles.PatternFill(start_color=color_obj, end_color=color_obj, fill_type="solid")

# Save new file
wb.save(OUTPUT_FILE)